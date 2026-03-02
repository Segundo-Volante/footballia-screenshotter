from pathlib import Path

import openpyxl
import pandas as pd

from backend.utils import logger


class ExcelManager:
    EXPECTED_COLUMNS = [
        "MD", "Date", "H/A", "Opponent", "Score", "Result",
        "Starting XI", "Substitutes (On)", "Goal Scorers",
        "Cards / Notes", "Referee",
    ]
    URL_COLUMN = "Footballia URL"

    def __init__(self, filepath: str):
        self.filepath = Path(filepath)
        if not self.filepath.exists():
            raise FileNotFoundError(f"Excel file not found: {self.filepath}")
        self.add_url_column_if_missing()

    def add_url_column_if_missing(self) -> None:
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb["Match Data"]

        headers = [cell.value for cell in ws[1]]
        if self.URL_COLUMN not in headers:
            col_idx = len(headers) + 1
            ws.cell(row=1, column=col_idx, value=self.URL_COLUMN)
            wb.save(self.filepath)
            logger.info(f"Added '{self.URL_COLUMN}' column to Excel")
        wb.close()

    def get_all_matches(self) -> list[dict]:
        df = pd.read_excel(self.filepath, sheet_name="Match Data")

        # Normalise column names for lookup
        col_map = {c: c for c in df.columns}

        matches = []
        for _, row in df.iterrows():
            md = row.get("MD", "")
            date_val = row.get("Date", "")
            if pd.isna(md) or md == "":
                continue

            # Format date
            if pd.notna(date_val):
                if hasattr(date_val, "strftime"):
                    date_str = date_val.strftime("%Y-%m-%d")
                else:
                    date_str = str(date_val)
            else:
                date_str = ""

            score = str(row.get("Score", "")) if pd.notna(row.get("Score", "")) else ""
            result = str(row.get("Result", "")) if pd.notna(row.get("Result", "")) else ""
            url = str(row.get(self.URL_COLUMN, "")) if pd.notna(row.get(self.URL_COLUMN, "")) else ""

            matches.append({
                "md": int(md) if pd.notna(md) else 0,
                "date": date_str,
                "home_away": str(row.get("H/A", "")) if pd.notna(row.get("H/A", "")) else "",
                "opponent": str(row.get("Opponent", "")) if pd.notna(row.get("Opponent", "")) else "",
                "score": score,
                "result": result,
                "starting_xi": str(row.get("Starting XI", "")) if pd.notna(row.get("Starting XI", "")) else "",
                "substitutes": str(row.get("Substitutes (On)", "")) if pd.notna(row.get("Substitutes (On)", "")) else "",
                "goal_scorers": str(row.get("Goal Scorers", "")) if pd.notna(row.get("Goal Scorers", "")) else "",
                "cards_notes": str(row.get("Cards / Notes", "")) if pd.notna(row.get("Cards / Notes", "")) else "",
                "referee": str(row.get("Referee", "")) if pd.notna(row.get("Referee", "")) else "",
                "footballia_url": url if url and url != "nan" else "",
            })

        return matches
